import datetime
import re

import requests
from bs4 import BeautifulSoup
import openpyxl
import time
import openpyxl
import pandas as pd
from pyautogui import size
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import subprocess
import shutil
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from bs4 import BeautifulSoup
import time
import datetime
import pyautogui
import pyperclip
import csv
import sys
import os
import math
import requests
import re
import random
import chromedriver_autoinstaller
from PyQt5.QtWidgets import QWidget, QApplication, QTreeView, QFileSystemModel, QVBoxLayout, QPushButton, QInputDialog, \
    QLineEdit, QMainWindow, QMessageBox, QFileDialog
from PyQt5.QtCore import QCoreApplication
from selenium.webdriver import ActionChains
from datetime import datetime, date, timedelta
import numpy
import datetime
# from window import Ui_MainWindow
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *




def chrome_browser(url):
    chrome_ver = chromedriver_autoinstaller.get_chrome_version().split('.')[0]  # 크롬 버전을 확인한다.
    driver_path = f'./{chrome_ver}/chromedriver.exe'
    if os.path.exists(driver_path):
        print(f"chromedriver is installed: {driver_path}")  # 있는 버전을 쓴다.
    else:
        print(f"install the chrome driver(ver: {chrome_ver})")  # 크롬을 최신 버전으로 설치한다.
        chromedriver_autoinstaller.install(True)

    options = webdriver.ChromeOptions()  # 크롬 옵션을 추가한다.
    # options.add_argument('headless')
    options.add_experimental_option("detach", True)  # 크롬 안 꺼지는 옵션 추가
    options.add_experimental_option("excludeSwitches", ["enable-logging"])  # 크롬 안 꺼지는 옵션 추가

    browser = webdriver.Chrome(driver_path, options=options)  # 크롬 드라이버를 할당
    browser.get(url)
    browser.maximize_window()
    browser.implicitly_wait(3)
    return browser


def load_excel(fname):
    wb = openpyxl.load_workbook(fname,data_only=True)
    ws = wb.active
    no_row = ws.max_row
    print("행갯수:", no_row)
    data_list = []
    for i in range(2, no_row + 1):
        name = ws.cell(row=i, column=5).value.replace("  "," ").replace("   "," ").replace("   "," ")
        if name == "" or name == None:
            print('데이타 더 이상 없음')
            break
        data_list.append(name)
    print(data_list)
    return data_list

end_flag=False


fname = 'list.xlsx'
name_list=load_excel(fname)


# count_flag=False


count=0
while True:
    try:
        url='https://www.esmplus.com/'
        browser=chrome_browser(url)

        id='moonstylecar'
        pw='8823msc##'
        btn_mall=browser.find_element(By.CSS_SELECTOR,'#rdoSiteSelect')
        browser.execute_script("arguments[0].click();", btn_mall)  #
        time.sleep(0.5)
        input_id=browser.find_element(By.CSS_SELECTOR,'#SiteId')
        input_id.send_keys(id)
        time.sleep(0.5)
        input_pw=browser.find_element(By.CSS_SELECTOR,'#SitePassword')
        input_pw.send_keys(pw)
        time.sleep(0.5)
        btn_login=browser.find_element(By.CSS_SELECTOR,'#btnSiteLogOn')
        browser.execute_script("arguments[0].click();", btn_login)  #
        time.sleep(1)

        while True:
            window_len=len(browser.window_handles)
            if window_len>=2:
                browser.switch_to.window(browser.window_handles[-1])
                browser.close()
                browser.switch_to.window(browser.window_handles[0])
            else:
                print("창 닫기 모두 완료")
                break

        btn_register_1=browser.find_element(By.CSS_SELECTOR,'#TDM001 > a')
        browser.execute_script("arguments[0].click();", btn_register_1)  #
        time.sleep(0.5)

        btn_register_2=browser.find_element(By.CSS_SELECTOR,'#TDM396 > a')
        browser.execute_script("arguments[0].click();", btn_register_2)  #
        time.sleep(1)


        # name_list에 이름들이 들어있다.



        no_name=len(name_list)
        register_count=0
        while True:
            register_count=register_count+1
            iframe_inner = browser.find_elements(By.CLASS_NAME, 'ifm_contents')[1]
            browser.switch_to.frame(iframe_inner)  # 프레임 이동

            gmarket_items = browser.find_element(By.CSS_SELECTOR,
                                                 '#divAcumlSum > table > tbody > tr:nth-child(2) > td.fir.aandg > a')

            browser.execute_script("arguments[0].click();", gmarket_items)  #
            time.sleep(2)

            last_row=browser.find_element(By.CLASS_NAME,'data_table').find_elements(By.TAG_NAME,'tr')[1]
            table_name=last_row.find_elements(By.TAG_NAME,'td')[7].text.strip()
            print("table_name:",table_name)


            for index_name,name in enumerate(name_list):
                if table_name==name.strip().replace(",",""):
                    count=index_name+1
                    print("{}번째 까지 등록 완료함".format(index_name))
                    # count_flag=True
                    break




            print("등록명:",name_list[count])
            btn_copy=last_row.find_elements(By.TAG_NAME,'a')[3]
            browser.execute_script("arguments[0].click();", btn_copy)  #
            time.sleep(3)

            browser.switch_to.default_content() # 원래대로 돌아가기
            iframe_inner=browser.find_elements(By.CLASS_NAME,'ifm_contents')[-1]
            browser.switch_to.frame(iframe_inner) #프레임 이동


            soup=BeautifulSoup(browser.page_source,'lxml')
            # print(soup.prettify())
            input_title=browser.find_element(By.CSS_SELECTOR,'#txtGoodsNameSearch')
            input_title.click()
            time.sleep(0.2)
            ActionChains(browser).key_down(Keys.CONTROL).send_keys('a').key_up(Keys.CONTROL).perform()
            time.sleep(0.2)
            ActionChains(browser).send_keys(Keys.DELETE).perform()
            time.sleep(0.2)
            ActionChains(browser).send_keys(name_list[count]).perform()

            btn_next=browser.find_element(By.CSS_SELECTOR,'#contents > div.syi_menu_control > a.button-step.button-step-next')
            browser.execute_script("arguments[0].click();", btn_next)  #

            additional_image=browser.find_element(By.CSS_SELECTOR,'#chkAdditionalImageVisable')
            browser.execute_script("arguments[0].click();", additional_image)  #
            time.sleep(0.2)

            file_path=r'D:\PythonProjects\069.gmarket_auto_register\photo'
            file_list = os.listdir(file_path)
            print(file_list)

            file_length=len(file_list)
            for file_index,file_elem in enumerate(file_list):
                while True:
                    div_box=browser.find_elements(By.CLASS_NAME,'img_box')[file_index]
                    if file_index==0:
                        btn_image_1=browser.find_element(By.CSS_SELECTOR,'#ulImageArea > li.ibox.fir > div > form:nth-child({}) > span.reg_btn_full > input'.format(file_index+1))
                    else:
                        btn_image_1 = browser.find_element(By.CSS_SELECTOR,'#ulImageArea > li:nth-child({}) > div > form:nth-child(1) > span.reg_btn_full > input'.format(file_index+1))
                    browser.execute_script("arguments[0].click();", btn_image_1)  #
                    time.sleep(0.5)
                    file_elem=file_elem.replace(".png","").replace(".jpg","")
                    pyperclip.copy(file_path+'\\'+file_elem)
                    pyautogui.hotkey('ctrl','v')
                    time.sleep(0.5)
                    pyautogui.hotkey('enter')
                    time.sleep(0.5)


                    img_tag=div_box.find_element(By.CLASS_NAME,'target').get_attribute('src')
                    print('img_tag:',img_tag)
                    if img_tag==None:
                        print("에러로 반복!")
                    else:
                        break

            btn_next=browser.find_element(By.CSS_SELECTOR,'#contents > div.syi_menu_control > a.button-step.button-step-next')
            browser.execute_script("arguments[0].click();", btn_next)  #
            browser.implicitly_wait(3)

            btn_no_abroad=browser.find_element(By.CSS_SELECTOR,'#SingleGoodsApp > div:nth-child(3) > div > table > tbody > tr.item.item_oversea-agree > td > div > div > div > p > label > input')
            browser.execute_script("arguments[0].click();", btn_no_abroad)  #
            browser.implicitly_wait(3)

            btn_next=browser.find_element(By.CSS_SELECTOR,'#contents > div.syi_menu_control > a.button-step.button-step-next')
            browser.execute_script("arguments[0].click();", btn_next)  #
            browser.implicitly_wait(3)

            btn_register=browser.find_element(By.CSS_SELECTOR,'#contents > div.syi_menu_wrap > div > div > a')
            browser.execute_script("arguments[0].click();", btn_register)  #
            browser.implicitly_wait(3)
            browser.switch_to.window(browser.window_handles[-1])

            while True:
                try:
                    label=browser.find_element(By.CSS_SELECTOR,'#lblConfirmForGoodsName').get_attribute('class')
                    print('label:',label)
                    if label.find("is-checked")>=0:
                        print("체크완료1")
                        break
                    btn_confirm1=browser.find_element(By.CSS_SELECTOR,'#lblConfirmForGoodsName')
                    # browser.execute_script("argument/s[0].click();", btn_confirm1)  #
                    btn_confirm1.click()
                    time.sleep(0.5)
                except:
                    print("인식에러1")

            while True:
                try:
                    label = browser.find_element(By.CSS_SELECTOR,'#lbConfirmForGoodsImage').get_attribute('class')
                    if label.find("is-checked") >= 0:
                        print("체크완료2")
                        break
                    btn_confirm2=browser.find_element(By.CSS_SELECTOR,'#lbConfirmForGoodsImage')
                    # browser.execute_script("arguments[0].click();", btn_confirm2)  #
                    btn_confirm2.click()
                    time.sleep(0.5)
                except:
                    print('인식에러2')

            btn_confirm3=browser.find_element(By.CSS_SELECTOR,'#btnConfirm')
            browser.execute_script("arguments[0].click();", btn_confirm3)  #
            time.sleep(1)

            browser.switch_to.window(browser.window_handles[0])

            iframe_inner = browser.find_elements(By.CLASS_NAME, 'ifm_contents')[-1]
            browser.switch_to.frame(iframe_inner)  # 프레임 이동
            check_complete=0
            while True:
                soup=BeautifulSoup(browser.page_source,'lxml')
                # print(soup.prettify())
                is_complete=len(soup.find_all('div',attrs={'class':'product_complete_group'}))
                print("is_complete:",is_complete,'check_complete:',check_complete)
                if is_complete>=1:
                    print("상품등록완료")
                    browser.switch_to.default_content()  # 원래대로 돌아가기
                    break
                if check_complete>=200:
                    print("완료된듯")
                    browser.switch_to.default_content()  # 원래대로 돌아가기
                    break
                check_complete=check_complete+1

                time.sleep(1)

            btn_close=browser.find_element(By.CSS_SELECTOR,'#tTDM411 > span > a')
            browser.execute_script("arguments[0].click();", btn_close)  #
            print("count:",count,"len_name_list:",len(name_list))
            if count==len(name_list)-1:
                print('등록완료!')
                end_flag=True
                break
            if register_count>=100:
                while True:
                    try:
                        browser.switch_to.window(browser.window_handles[-1])
                        browser.close()
                    except:
                        print("전부닫기완료")
                        break
                break

        if end_flag==True:
            print('등록완료!!')
            break
    except:
        while True:
            try:
                browser.switch_to.window(browser.window_handles[-1])
                browser.close()
            except:
                print("전부닫기완료")
                break


print("종료")


